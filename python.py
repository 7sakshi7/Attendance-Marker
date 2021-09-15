from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import random
import os
import ast
from functools import partial
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime


class create_profile_page(Toplevel):
    def __init__(self, master):
        super().__init__(master=master)
        self.title('Create Profile')
        self.geometry('800x500+400+0')

        self.username, self.department_name, self.city_name = StringVar(), StringVar(), StringVar()

        frame = LabelFrame(self, bg="teal")

        top_title = Label(frame, bd=2, relief=RIDGE, text="CREATE PROFILE", font=(
            "times new roman", 20), bg="#ff9900", pady=20)
        top_title.pack(side=TOP, fill=X)

        self.profile = LabelFrame(frame, bd=4, padx=100, pady=30, bg="#ff5959")

        self.name = Label(self.profile, font=("arial", 12),
                          text="Your Name  ", padx=3, pady=6, bg="#ff5959")
        self.name.grid(row=0, column=0)
        self.name_text = Entry(
            self.profile, textvariable=self.username, font=("arial", 12), width=35)
        self.name_text.grid(row=0, column=1)

        self.department = Label(self.profile, font=(
            "arial", 12), text="Department  ", padx=3, pady=6, bg="#ff5959")
        self.department.grid(row=1, column=0)
        self.department_text = Entry(
            self.profile, textvariable=self.department_name, font=("arial", 12), width=35)
        self.department_text.grid(row=1, column=1)

        self.city = Label(self.profile, font=("arial", 12),
                          text="Your City  ", padx=3, pady=6, bg="#ff5959")
        self.city.grid(row=2, column=0)
        self.city_text = Entry(
            self.profile, textvariable=self.city_name, font=("arial", 12), width=35)
        self.city_text.grid(row=2, column=1)

        generate = Button(self.profile, text="Generate Unique Id", command=self.user_info, bg="green", fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"), width=28)
        generate.grid(row=4, column=1)

        self.profile.pack()

        instructions = Label(frame, bg="teal", fg="white", text=''' 
        1. Enter Your Name,Department in which you are in and the city you belong.
        2. Click On Generate Unique id.
        3. Id will be Generated.
        4. Click On Home Button to go back to Main Screen
        ''')
        instructions.pack()
        home = Button(frame, text="Home", bg="green", fg="white", padx=5, pady=5, font=(
            "arial", 12), width=18, command=self.destroy)
        home.pack()

        frame.place(x=0, y=0, width=800, height=500)

    def generate_unique_id(self):
        return random.randint(1000, 9999)

    def clear_fields(self):
        self.name_text.delete(0, 'end')
        self.department_text.delete(0, 'end')
        self.city_text.delete(0, 'end')

    def user_info(self):
        if self.username.get() == "" or self.department_name.get() == "" or self.city_name == "":
            messagebox.showerror('Error', 'All the field are required')
        else:
            id = self.generate_unique_id()
            if not os.path.isfile('id.xlsx'):
                wb = Workbook()
                ws = wb.active
                ws.title = "Information"
                headings = ["Id", "Name", "Department", "City"]
                ws.append(headings)
                data = [id, self.username.get(), self.department_name.get(),
                        self.city_name.get()]
                ws.append(data)
                wns = wb.create_sheet("Attendance")
                wns.append(["Date", "Id", "Lecture-1", "Lecture-2", "Lecture-3",
                           "Lecture-4", "Lecture-5", "Lecture-6", "Lecture-7"])
                wb.save('id.xlsx')
            else:
                wb = load_workbook('id.xlsx')
                ws = wb.active
                col = 1
                while(ws['A'+str(col)].value != None):
                    print(type(id), type(ws['A'+str(col)].value))
                    if(str(id) == ws['A'+str(col)].value):
                        id = self.generate_unique_id()
                    col += 1

                data = [id, self.username.get(), self.department_name.get(),
                        self.city_name.get()]
                ws.append(data)
                wb.save('id.xlsx')

            self.clear_fields()
            id_label = Label(self.profile, text=id, padx=20, pady=10)
            id_label.grid(row=6, column=1)


class mark_attendance_screen(Toplevel):
    def __init__(self, master):
        super().__init__(master=master)
        # title
        self.title('Mark Attendance')
        self.geometry('500x500+400+0')

        # variable for taking usre id input
        user_attendance_id_text = StringVar()

        # whole frame
        frame = LabelFrame(self, bg="#f75648")

        # title inside frame
        top_title = Label(frame, bd=2, relief=RIDGE, text="MARK YOUR ATTENDANCE", font=(
            "times new roman", 20), bg="#ff9900", pady=20)
        top_title.pack(side=TOP, fill=X)

        # frame for taking input
        self.mark_label = LabelFrame(
            frame, bd=3, padx=10, pady=10, bg="#f75648")

        user_attendance = Label(self.mark_label, text="Enter Your Unique Id : ", font=(
            "arial", 12), padx=5, pady=5,)
        user_attendance.place(x=10, y=20, width=200)
        self.user_attendance_text = Entry(self.mark_label, textvariable=user_attendance_id_text, font=(
            "arial", 12), width=20, relief=RIDGE)
        self.user_attendance_text.place(x=220, y=25, width=100)

        mark_button = Button(self.mark_label, text="MARK", command=partial(self.add_attendance, user_attendance_id_text), bg="green", fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"))
        mark_button.pack(padx=30, pady=60)

        home = Button(self.mark_label, text="HOME", command=self.destroy, bg="green", fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"))
        home.place(x=350, y=100)

        self.mark_label.place(x=30, y=90, width=450, height=200)

        frame.place(x=0, y=0, width=500, height=500)

    def add_attendance(self, user_id):
        safe = True
        try:
            int(user_id.get())
        except ValueError:
            messagebox.showerror('Error', 'Id doesnot exist')
            return
        if not os.path.isfile('id.xlsx') or user_id.get() == "":
            messagebox.showerror('Error', 'Id doesnot exist')
        else:
            found = False
            wb = load_workbook('id.xlsx')
            wbs = wb['Information']
            col = 1
            while(found == False):
                if(wbs['A'+str(col)].value == int(user_id.get())):
                    found = True
                    break
                if(wbs['A'+str(col)].value == None):
                    break
                col += 1

            if found == False:
                messagebox.showerror('Error', 'Id doesnot exist')
            else:
                ws = wb['Attendance']
                now = datetime.now()
                date_string = now.strftime("%d/%m/%Y")
                col = 1
                found = False
                while(found == False):
                    print(type(ws['B'+str(col)].value), type(user_id.get()))
                    print((ws['B'+str(col)].value), (user_id.get()))
                    if(ws['B'+str(col)].value == user_id.get()):
                        found = True
                        break
                    if(ws['B'+str(col)].value == None):
                        break
                    col += 1
                # ws['A'+str(col)] = date_string
                # wb.save('id.xlsx')
                # ws['B'+str(col)] = str(user_id.get())
                print(ws['B'+str(col)].value)

                time_string = now.strftime("%H:%M:%S")
                print(time_string)
                # ws.insert_rows(col-1)
                id_value = (user_id.get())
                if(time_string >= "08:50:00" and time_string < "09:50:00"):
                    if(ws['C'+str(col)].value == None or ws['C'+str(col)].value == 0):
                        dt = [date_string, id_value,
                              time_string, 0, 0, 0, 0, 0, 0]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1
                        # ws.append()
                        # ws['C'+str(col)] = time_string
                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')

                elif(time_string >= "09:50:00" and time_string < "10:50:00"):
                    if(ws['D'+str(col)].value == None or ws['D'+str(col)].value == 0):
                        dt = [date_string, id_value, 0,
                              time_string, 0, 0, 0, 0, 0]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1
                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')

                elif(time_string >= "10:50:00" and time_string < "11:50:00"):
                    if(ws['E'+str(col)].value == None or ws['E'+str(col)].value == 0):
                        dt = [date_string, id_value, 0,
                              0, time_string, 0, 0, 0, 0]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1
                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')

                elif(time_string >= "11:50:00" and time_string < "12:50:00"):
                    if(ws['F'+str(col)].value == None or ws['F'+str(col)].value == 0):
                        dt = [date_string, id_value, 0,
                              0, 0, time_string, 0, 0, 0]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1

                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')

                elif(time_string >= "01:40:00" and time_string < "02:30:00"):
                    if(ws['G'+str(col)].value == None or ws['G'+str(col)].value == 0):
                        dt = [date_string, id_value, 0,
                              0, 0, 0, time_string, 0, 0]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1

                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')

                elif(time_string >= "02:30:00" and time_string < "03:20:00"):
                    if(ws['H'+str(col)].value == None or ws['H'+str(col)].value == 0):
                        dt = [date_string, id_value, 0,
                              0, 0, 0, 0, time_string, 0]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1
                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')

                elif(time_string >= "03:20:00"):
                    if(ws['I'+str(col)].value == None or ws['I'+str(col)].value == 0):
                        dt = [date_string, id_value, 0,
                              0, 0, 0, 0, 0, time_string]
                        letter = 1
                        for details in dt:
                            ws[get_column_letter(letter)+str(col)] = details
                            letter += 1
                    else:
                        messagebox.showinfo(
                            'Done', 'Attendance Already Marked for this Lecture')
                # ws.insert_rows(col-1)
                wb.save('id.xlsx')
                # ws.append([date_string,id,])
                marked = Label(self.mark_label, text="Attendance Marked", bd=2, relief=RIDGE, font=(
                    "times new roman", 12), padx=30, pady=10)
                marked.place(x=10, y=100, width=150, height=50)
                self.user_attendance_text.delete(0, 'end')


class update_profile_page(Toplevel):
    def __init__(self, master):
        super().__init__(master=master)
        # title
        self.title('Update Profile')
        self.geometry('700x500+400+0')

        # variables
        self.username, self.department_name, self.city_name, self.user_id = StringVar(
        ), StringVar(), StringVar(), StringVar()

        # whole frame
        frame = LabelFrame(self, bg="#019fb8")
        top_title = Label(frame, bd=2, relief=RIDGE, text="UPDATE PROFILE", font=(
            "times new roman", 20), bg="#ff9900", pady=20)
        top_title.pack(side=TOP, fill=X)

        self.profile = LabelFrame(frame, bd=4, padx=100, pady=50, bg="#ff5959")

        self.enter_id = Label(self.profile, font=("arial", 12),
                              text="Your Id  ", padx=3, pady=6, bg="#ff5959")
        self.enter_id.grid(row=0, column=0)
        self.enter_id_text = Entry(
            self.profile, textvariable=self.user_id, font=("arial", 12), width=35)
        self.enter_id_text.grid(row=0, column=1)

        self.name = Label(self.profile, font=("arial", 12),
                          text="Your Name  ", padx=3, pady=6, bg="#ff5959")
        self.name.grid(row=1, column=0)
        self.name_text = Entry(
            self.profile, textvariable=self.username, font=("arial", 12), width=35)
        self.name_text.grid(row=1, column=1)

        self.department = Label(self.profile, font=(
            "arial", 12), text="Department  ", padx=3, pady=6, bg="#ff5959")
        self.department.grid(row=2, column=0)
        self.department_text = Entry(
            self.profile, textvariable=self.department_name, font=("arial", 12), width=35)
        self.department_text.grid(row=2, column=1)

        self.city = Label(self.profile, font=("arial", 12),
                          text="Your City  ", padx=3, pady=6, bg="#ff5959")
        self.city.grid(row=4, column=0)
        self.city_text = Entry(
            self.profile, textvariable=self.city_name, font=("arial", 12), width=35)
        self.city_text.grid(row=4, column=1)

        update = Button(self.profile, text="Update", bg="green", command=partial(self.update_data_to_database, self.user_id, self.username, self.department_name, self.city_name), fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"), width=15)
        update.grid(row=6, columnspan=1)

        exit = Button(self.profile, text="Exit", bg="green", command=self.destroy, fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"), width=15)
        exit.grid(row=6, column=1)

        self.profile.pack()

        frame.place(x=0, y=0, width=700, height=500)

    def update_data_to_database(self, user_id, username, department, city):
        try:
            int(user_id.get())
        except ValueError:
            messagebox.showerror('Error', 'Id doesnot exist')
            return
        if not os.path.isfile('id.xlsx') or user_id.get() == "":
            messagebox.showerror('Error', 'Id doesnot exist')
        else:
            found = False
            found = False
            wb = load_workbook('id.xlsx')
            wbs = wb['Information']
            col = 1
            while(found == False):
                print(type(wbs['A'+str(col)].value))
                if(wbs['A'+str(col)].value == int(user_id.get())):
                    found = True
                    break
                if(wbs['A'+str(col)].value == None):
                    break
                col += 1

            if found == False:
                messagebox.showerror('Error', 'Id doesnot exist')
            else:
                if(self.username.get() == ""):
                    pass
                else:
                    wbs['B'+str(col)] = self.username.get()

                if(self.department_name.get() == ""):
                    pass
                else:
                    wbs['C'+str(col)] = self.department_name.get()

                if(self.city_name.get() == ""):
                    pass
                else:
                    wbs['D'+str(col)] = self.city_name.get()

            wb.save('id.xlsx')

            self.enter_id_text.delete(0, 'end')
            self.name_text.delete(0, 'end')
            self.department_text.delete(0, 'end')
            self.city_text.delete(0, 'end')
            updated = Label(self.profile, text="Updated", bg="green", fg="white", padx=10, pady=10, font=(
                "arial", 12, "bold"), width=15)
            updated.grid(row=7, columnspan=2)


class view_attendance_screen(Toplevel):
    def __init__(self, master):
        super().__init__(master=master)
        # title
        self.title('View Attendance')
        self.geometry('500x750+400+0')

        # variable for taking usre id input
        self.user_attendance_id_text = StringVar()

        # whole frame
        frame = LabelFrame(self, bg="#019fb8")

        # title inside frame
        top_title = Label(frame, bd=2, relief=RIDGE, text="VIEW YOUR ATTENDANCE", font=(
            "times new roman", 20), bg="#f87604", pady=20)
        top_title.pack(side=TOP, fill=X)

        # frame for taking input id
        self.mark_label = LabelFrame(
            frame, bd=3, padx=10, pady=10, bg="#9d196c")

        user_attendance = Label(self.mark_label, text="Enter Your Unique Id : ", font=(
            "arial", 12), padx=5, pady=5,)
        user_attendance.place(x=10, y=20, width=200)
        self.user_attendance_text = Entry(self.mark_label, textvariable=self.user_attendance_id_text, font=(
            "arial", 12), width=20, relief=RIDGE)
        self.user_attendance_text.place(x=220, y=25, width=100)

        view_button = Button(self.mark_label, text="VIEW", bg="#057790", fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"), command=partial(self.view_attendance, self.user_attendance_id_text))
        view_button.pack(padx=30, pady=60)

        home = Button(self.mark_label, text="HOME", command=self.destroy, bg="#057790", fg="white", padx=10, pady=10, font=(
            "arial", 12, "bold"))
        home.place(x=320, y=550)

        self.mark_label.place(x=30, y=90, width=450, height=650)

        frame.place(x=0, y=0, width=500, height=750)

    def view_attendance(self, user_id):
        try:
            int(user_id.get())
        except ValueError:
            messagebox.showerror('Error', 'Id doesnot exist')
            return
        if not os.path.isfile('id.xlsx') or user_id.get() == "":
            messagebox.showerror('Error', 'Id doesnot exist')
        else:
            found = False
            wb = load_workbook('id.xlsx')
            wbs = wb['Information']
            col_info = 1
            while(found == False):
                if(wbs['A'+str(col_info)].value == int(user_id.get())):
                    found = True
                    break
                if(wbs['A'+str(col_info)].value == None):
                    break
                col_info += 1
            ws = wb["Attendance"]
            col = 1
            found_att = False
            attendance = [0, 0, 0, 0, 0, 0, 0]
            while(found_att == False):
                if(ws['B'+str(col)].value == str(user_id.get())):
                    lec = 3
                    for i in range(7):
                        if(ws[get_column_letter(lec)+str(col)].value == 0):
                            pass
                        else:
                            attendance[i] += 1
                        lec += 1

                if(ws['A'+str(col)].value == None):
                    break
                col += 1
            if found == False:
                messagebox.showerror('Error', 'Id doesnot exist')
            else:
                # name
                name = Label(self.mark_label, font=(
                    "arial", 12), text="Your Name  ", padx=3, pady=3, bg="#9d196c", fg="white")
                name.place(x=0, y=130, width=100, height=50)

                name_text = Label(
                    self.mark_label, text=wbs['B'+str(col_info)].value, font=("arial", 12), bg="#9d196c", fg="white")
                name_text.place(x=100, y=130, width=100, height=50)

                # department
                department = Label(self.mark_label, font=(
                    "arial", 12), text="Department  ", padx=3, pady=3, bg="#9d196c", fg="white")
                department.place(x=0, y=180, width=100, height=50)
                department_text = Label(
                    self.mark_label, text=wbs['C'+str(col_info)].value, font=("arial", 12), bg="#9d196c", fg="white")
                department_text.place(
                    x=100, y=180, width=100, height=50)

                # city
                city = Label(self.mark_label, font=("arial", 12),
                             text="Your City  ", padx=3, pady=3, bg="#9d196c", fg="white")
                city.place(x=0, y=230, width=100, height=50)
                city_text = Label(
                    self.mark_label, text=wbs['D'+str(col_info)].value, font=("arial", 12), bg="#9d196c", fg="white")
                city_text.place(x=100, y=230, width=100, height=50)

                # attendance
                # lecture 1
                lect1 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-1", padx=0, pady=3, bg="#9d196c", fg="white")
                lect1.place(x=0, y=270, width=100, height=80)
                lect1_text = Label(
                    self.mark_label, text=attendance[0], font=("arial", 12), bg="#9d196c", fg="white")
                lect1_text.place(
                    x=100, y=280, width=100, height=50)

                # lecture 2
                lect2 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-2", padx=0, pady=3, bg="#9d196c", fg="white")
                lect2.place(x=0, y=320, width=100, height=80)
                lect2_text = Label(
                    self.mark_label, text=attendance[1], font=("arial", 12), bg="#9d196c", fg="white")
                lect2_text.place(
                    x=100, y=330, width=100, height=50)

                # lecture 3
                lect3 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-3", padx=0, pady=3, bg="#9d196c", fg="white")
                lect3.place(x=0, y=370, width=100, height=80)
                lect3_text = Label(
                    self.mark_label, text=attendance[2], font=("arial", 12), bg="#9d196c", fg="white")
                lect3_text.place(
                    x=100, y=380, width=100, height=50)

                # lecture 4
                lect4 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-4", padx=0, pady=3, bg="#9d196c", fg="white")
                lect4.place(x=0, y=420, width=100, height=80)
                lect4_text = Label(
                    self.mark_label, text=attendance[3], font=("arial", 12), bg="#9d196c", fg="white")
                lect4_text.place(
                    x=100, y=430, width=100, height=50)

                # lecture 5
                lect5 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-5", padx=0, pady=3, bg="#9d196c", fg="white")
                lect5.place(x=0, y=470, width=100, height=80)
                lect5_text = Label(
                    self.mark_label, text=attendance[4], font=("arial", 12), bg="#9d196c", fg="white")
                lect5_text.place(
                    x=100, y=480, width=100, height=50)

                # lecture 6
                lect6 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-6", padx=0, pady=3, bg="#9d196c", fg="white")
                lect6.place(x=0, y=520, width=100, height=80)
                lect6_text = Label(
                    self.mark_label, text=attendance[5], font=("arial", 12), bg="#9d196c", fg="white")
                lect6_text.place(
                    x=100, y=530, width=100, height=50)

                # lecture 7
                lect7 = Label(self.mark_label, font=("arial", 12),
                              text="Lecture-7", padx=0, pady=3, bg="#9d196c", fg="white")
                lect7.place(x=0, y=570, width=100, height=80)
                lect7_text = Label(
                    self.mark_label, text=attendance[6], font=("arial", 12), bg="#9d196c", fg="white")
                lect7_text.place(
                    x=100, y=580, width=100, height=50)


class MainScreen():
    def __init__(self, root):
        self.root = root
        self.root.title('Main Screen')
        self.root.geometry('800x530+200+0')

        label_title = Label(self.root, bd=10, relief=RIDGE, text="Heyy Welcome !!",
                            fg="white", bg="purple", font=("times new roman", 30, "bold"))
        label_title.pack(side=TOP, fill=X)

        self.options = LabelFrame(self.root, text='Options', relief=RIDGE, padx=10, font=(
            "arial", 12, "bold"), bd=5)
        self.options.place(x=50, y=90, width=650, height=430)
        self.create_buttons()

    def create_buttons(self):
        mark_attendance = Button(self.options, text="Mark Attendance", command=self.mark, bg="green", fg="white", padx=10, pady=20, font=(
            "arial", 12, "bold"), width=28)
        mark_attendance.place(x=150, y=10, height=50)

        create_profile = Button(self.options, text="Create Profile", command=self.create, bg="green", fg="white", padx=10, pady=20, font=(
            "arial", 12, "bold"), width=28)

        create_profile.place(x=150, y=90, height=50)

        view_attendance = Button(self.options, text="View Attendance", command=self.view, bg="green", fg="white", padx=10, pady=20, font=(
            "arial", 12, "bold"), width=28)
        view_attendance.place(x=150, y=170, height=50)

        update = Button(self.options, text="Update Profile", command=self.update_profile, bg="green", fg="white", padx=10, pady=20, font=(
            "arial", 12, "bold"), width=28)
        update.place(x=150, y=250, height=50)

        quit = Button(self.options, text="Quit", bg="green", command=root.destroy, fg="white", padx=10, pady=20, font=(
            "arial", 12, "bold"), width=28)
        quit.place(x=150, y=330, height=50)

    def create(self):
        create_profile_page(self.root)

    def mark(self):
        mark_attendance_screen(self.root)

    def view(self):
        view_attendance_screen(self.root)

    def update_profile(self):
        update_profile_page(self.root)


root = Tk()
obj = MainScreen(root)
root.mainloop()
